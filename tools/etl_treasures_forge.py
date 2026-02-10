#!/usr/bin/env python3
"""
ETL: 覇道DB_名宝_.xlsx → data-treasures-forge.js
名宝一覧シート + 鍛錬シートを読み込み、
鍛錬ランク別の技能レベルデータをJSファイルとして出力する。

使い方:
  python etl_treasures_forge.py 覇道DB_名宝_.xlsx data-treasures-forge.js

読込仕様:
  名宝一覧シート: 行1=ヘッダー, 行2〜=データ（209件）
  鍛錬シート:     行1=ヘッダー, 行2〜=データ（1,030行）
"""

import sys
import json
import unicodedata
from collections import defaultdict
from openpyxl import load_workbook


def nfkc(val):
    if isinstance(val, str):
        return unicodedata.normalize('NFKC', val).strip()
    return val


def safe_str(val):
    if val is None:
        return None
    s = nfkc(str(val))
    return s if s else None


def safe_int(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(float(nfkc(str(val))))
    except (ValueError, TypeError):
        return None


def read_treasure_list(wb):
    """名宝一覧シートを読み込む"""
    ws = wb['名宝一覧']
    treasures = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        t_id = safe_int(row[0])
        if t_id is None:
            continue

        treasures[t_id] = {
            'id': t_id,
            'name': safe_str(row[4]) if len(row) > 4 else None,   # 列5: 名称
            'type': safe_str(row[3]) if len(row) > 3 else None,   # 列4: 種類
            'related': safe_str(row[5]) if len(row) > 5 else None, # 列6: 関連武将
            'faction': safe_str(row[6]) if len(row) > 6 else None, # 列7: 関連1
            'source': safe_str(row[7]) if len(row) > 7 else None,  # 列8: 関連2
            'skills': [],
            'isUR': safe_str(row[11]) == '〇' if len(row) > 11 else False,
            'isKokou': safe_str(row[12]) == '〇' if len(row) > 12 else False,
        }

        # 技能名（列9-11）
        for i in range(3):
            skill = safe_str(row[8 + i]) if len(row) > 8 + i else None
            if skill and skill != '(技能データなし)':
                treasures[t_id]['skills'].append(skill)

    return treasures


def read_forge_data(wb):
    """鍛錬シートを読み込む"""
    ws = wb['鍛錬']
    forge_data = defaultdict(lambda: {'normal': {}, 'ur': {}})

    # ☆/★ランク列名（列5〜15: ☆0〜★10）
    rank_names = ['☆0', '☆1', '☆2', '☆3', '☆4', '☆5', '☆6', '☆7', '★8', '★9', '★10']

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        t_id = safe_int(row[0])
        stage = safe_str(row[2])  # 列3: 段階（通常/UR）
        skill_name = safe_str(row[3])  # 列4: 技能名

        if t_id is None or not stage or not skill_name:
            continue

        # 技能データなしの場合はスキップ
        if skill_name == '(技能データなし)':
            continue

        # ☆ランク別レベル
        levels = {}
        for i, rank_name in enumerate(rank_names):
            val = row[4 + i] if len(row) > 4 + i else None
            level = safe_int(val)
            if level is not None and level > 0:
                levels[rank_name] = level

        target = 'ur' if stage == 'UR' else 'normal'
        forge_data[t_id][target][skill_name] = levels

    return dict(forge_data)


def write_js(treasures, forge_data, output_path):
    """鍛錬データをJSファイルに書き出す"""

    # 名宝IDと鍛錬データを統合
    result = {}
    for t_id, treasure in treasures.items():
        forge = forge_data.get(t_id, {'normal': {}, 'ur': {}})

        entry = {
            'id': t_id,
            'name': treasure['name'],
            'type': treasure['type'],
            'related': treasure['related'],
            'skills': treasure['skills'],
            'isUR': treasure['isUR'],
        }

        # 鍛錬データ
        if forge['normal']:
            entry['forge'] = forge['normal']
        if forge['ur']:
            entry['forgeUR'] = forge['ur']

        result[treasure['name'] or str(t_id)] = entry

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('// data-treasures-forge.js — 名宝鍛錬データ（自動生成）\n')
        f.write('// 生成元: 覇道DB_名宝_.xlsx\n')
        f.write(f'// 名宝数: {len(result)}\n\n')
        f.write('const TREASURES_FORGE_DATA = ')
        f.write(json.dumps(result, ensure_ascii=False, indent=2))
        f.write(';\n')

    print(f'✅ {output_path} 生成完了')
    print(f'   名宝数: {len(result)}')
    forge_count = sum(1 for t in result.values() if 'forge' in t or 'forgeUR' in t)
    print(f'   鍛錬データあり: {forge_count}件')


def main():
    if len(sys.argv) < 3:
        print('使い方: python etl_treasures_forge.py <覇道DB_名宝_.xlsx> <出力先.js>')
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    print(f'📖 読込中: {input_path}')
    wb = load_workbook(input_path, read_only=True, data_only=True)

    if '名宝一覧' not in wb.sheetnames:
        print('❌ 名宝一覧 シートが見つかりません')
        sys.exit(1)
    if '鍛錬' not in wb.sheetnames:
        print('❌ 鍛錬 シートが見つかりません')
        sys.exit(1)

    treasures = read_treasure_list(wb)
    forge_data = read_forge_data(wb)
    wb.close()

    write_js(treasures, forge_data, output_path)


if __name__ == '__main__':
    main()
