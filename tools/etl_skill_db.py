#!/usr/bin/env python3
"""
ETL: 覇道DB.xlsx → data-skill-db.js
技能DBシート（Power Query統合ビュー）を読み込み、
SKILL_DB オブジェクトをJSファイルとして出力する。

使い方:
  python etl_skill_db.py 覇道DB.xlsx data-skill-db.js

読込仕様:
  - 行1: 情報行（スキップ）
  - 行2: テーブルヘッダー（39列）
  - 行3〜: データ（5,733行）
"""

import sys
import json
import unicodedata
from collections import defaultdict
from openpyxl import load_workbook

# === 定数 ===

# 共通列の列番号（1-indexed）
COL_MAP = {
    'ID': 1,
    '実装日': 2,
    '更新日時': 3,
    '名称': 4,
    'ﾚｱ': 5,
    '関連': 6,
    '種類1': 7,
    '種類2': 8,
    '効果': 9,
    '効果x': 10,
    '発動条件': 11,
    '発動条件2': 12,
    '発動タイミング': 13,
    '追加効果': 14,
    '確率': 15,
    '対象部隊数': 16,
    '効果時間': 17,
    '効果量': 18,
    # Ⅰ〜Ⅹ: 19〜28
    '説明文': 29,
    # 分析9列: 30〜38（39列シート）or 30〜38（38列シート）
}

LEVEL_COLS = {
    'Ⅰ': 19, 'Ⅱ': 20, 'Ⅲ': 21, 'Ⅳ': 22, 'Ⅴ': 23,
    'Ⅵ': 24, 'Ⅶ': 25, 'Ⅷ': 26, 'Ⅸ': 27, 'Ⅹ': 28,
}

# 分析9列（39列シートの場合: 列30-38、Name=列39）
ANALYSIS_COLS_39 = {
    '効果No': 30, '値種別': 31, '乗算変数': 32, '乗算対象': 33,
    '追加条件': 34, '条件倍率': 35, '上限値': 36, '発動回数': 37, '攻撃系統': 38,
}
# Name列: 39列シート→列39、38列シート→なし


def nfkc(val):
    """Unicode NFKC正規化"""
    if isinstance(val, str):
        return unicodedata.normalize('NFKC', val).strip()
    return val


def safe_str(val):
    """セル値を安全に文字列化"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return str(val) if val != 0 else None
    s = nfkc(str(val))
    return s if s else None


def safe_num(val):
    """セル値を安全に数値化"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    try:
        return float(nfkc(str(val)))
    except (ValueError, TypeError):
        return None


def read_skill_db_sheet(wb, sheet_name='技能DB'):
    """技能DBシートを読み込む"""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # 行3〜（行1=情報行, 行2=ヘッダー）

    # ヘッダー確認（行2）
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    total_cols = len([c for c in header_row if c is not None])
    has_name_col = total_cols >= 39

    skills = defaultdict(lambda: {'id': None, 'effects': [], 'owners': [], 'source': None})
    seen_names = set()

    for row_idx, row in enumerate(rows):
        if not row or len(row) < 18:
            continue

        skill_name = safe_str(row[COL_MAP['名称'] - 1])
        if not skill_name:
            continue

        skill_id = safe_num(row[COL_MAP['ID'] - 1])
        type2 = safe_str(row[COL_MAP['種類2'] - 1])

        # Nameの位置（39列→列39、38列→なし）
        source = None
        if has_name_col and len(row) >= 39:
            source = safe_str(row[38])  # 0-indexed: 列39→index 38

        entry = skills[skill_name]
        if entry['id'] is None and skill_id is not None:
            entry['id'] = int(skill_id)
        if source and not entry['source']:
            entry['source'] = source

        # 所持行の処理
        if type2 == '所持':
            rarity = safe_str(row[COL_MAP['ﾚｱ'] - 1])
            related = safe_str(row[COL_MAP['関連'] - 1])
            if related:
                entry['owners'].append({'name': related, 'rarity': rarity or ''})
            continue

        # 効果行の処理
        effect = {
            'effectNo': safe_num(row[ANALYSIS_COLS_39['効果No'] - 1]) if len(row) >= 30 else None,
            'type1': safe_str(row[COL_MAP['種類1'] - 1]),
            'type2': type2,
            'effect': safe_str(row[COL_MAP['効果'] - 1]),
            'effectX': safe_str(row[COL_MAP['効果x'] - 1]),
            'condition': safe_str(row[COL_MAP['発動条件'] - 1]),
            'condition2': safe_str(row[COL_MAP['発動条件2'] - 1]),
            'timing': safe_str(row[COL_MAP['発動タイミング'] - 1]),
            'extra': safe_str(row[COL_MAP['追加効果'] - 1]),
            'probability': safe_str(row[COL_MAP['確率'] - 1]),
            'targetCount': safe_str(row[COL_MAP['対象部隊数'] - 1]),
            'duration': safe_str(row[COL_MAP['効果時間'] - 1]),
            'amount': safe_str(row[COL_MAP['効果量'] - 1]),
            'levels': {},
            'description': safe_str(row[COL_MAP['説明文'] - 1]),
        }

        # Ⅰ〜Ⅹ のレベル値
        for lv_name, col in LEVEL_COLS.items():
            val = safe_num(row[col - 1]) if len(row) >= col else None
            if val is not None:
                effect['levels'][lv_name] = val

        # 分析9列
        if len(row) >= 38:
            effect['valueType'] = safe_str(row[ANALYSIS_COLS_39['値種別'] - 1])
            effect['multiVar'] = safe_str(row[ANALYSIS_COLS_39['乗算変数'] - 1])
            effect['multiTarget'] = safe_str(row[ANALYSIS_COLS_39['乗算対象'] - 1])
            effect['extraCondition'] = safe_str(row[ANALYSIS_COLS_39['追加条件'] - 1])
            effect['conditionRate'] = safe_num(row[ANALYSIS_COLS_39['条件倍率'] - 1])
            effect['capValue'] = safe_num(row[ANALYSIS_COLS_39['上限値'] - 1])
            effect['activationCount'] = safe_num(row[ANALYSIS_COLS_39['発動回数'] - 1])
            effect['attackType'] = safe_str(row[ANALYSIS_COLS_39['攻撃系統'] - 1])

        # effectNoがintの場合は整数化
        if effect['effectNo'] is not None:
            effect['effectNo'] = int(effect['effectNo'])
        if effect['conditionRate'] is not None and effect['conditionRate'] == int(effect['conditionRate']):
            effect['conditionRate'] = int(effect['conditionRate'])
        if effect['activationCount'] is not None and effect['activationCount'] == int(effect['activationCount']):
            effect['activationCount'] = int(effect['activationCount'])

        entry['effects'].append(effect)

    return dict(skills)


def write_js(skills, output_path):
    """SKILL_DB を JSファイルに書き出す"""

    def clean_none(obj):
        """None値をnullに、空dictを除去"""
        if isinstance(obj, dict):
            return {k: clean_none(v) for k, v in obj.items() if v is not None}
        if isinstance(obj, list):
            return [clean_none(v) for v in obj]
        return obj

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('// data-skill-db.js — 技能DB全量（自動生成）\n')
        f.write('// 生成元: 覇道DB.xlsx 技能DBシート\n')
        f.write(f'// 技能数: {len(skills)}\n')

        total_effects = sum(len(v['effects']) for v in skills.values())
        f.write(f'// 効果行: {total_effects}\n\n')
        f.write('const SKILL_DB = ')

        # JSON変換（読みやすいインデント）
        cleaned = {}
        for name, data in skills.items():
            cleaned[name] = clean_none(data)

        json_str = json.dumps(cleaned, ensure_ascii=False, indent=2)
        f.write(json_str)
        f.write(';\n')

    print(f'✅ {output_path} 生成完了')
    print(f'   技能数: {len(skills)}')
    print(f'   効果行: {total_effects}')


def main():
    if len(sys.argv) < 3:
        print('使い方: python etl_skill_db.py <覇道DB.xlsx> <出力先.js>')
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    print(f'📖 読込中: {input_path}')
    wb = load_workbook(input_path, read_only=True, data_only=True)

    if '技能DB' not in wb.sheetnames:
        print('❌ 技能DB シートが見つかりません')
        sys.exit(1)

    skills = read_skill_db_sheet(wb)
    wb.close()

    write_js(skills, output_path)


if __name__ == '__main__':
    main()
