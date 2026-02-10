#!/usr/bin/env python3
"""
バリデーション: 覇道DB データ整合性チェック

使い方:
  python validate_data.py 覇道DB.xlsx 覇道DB_武将_.xlsx 覇道DB_名宝_.xlsx

チェック項目:
  1. 武将IDの一意性（レアリティ横断）
  2. 武将技能→技能DBリンクの整合性
  3. 名宝技能→技能DBリンクの整合性
  4. 名宝ID衝突の検出
  5. 鍛錬データと名宝一覧の整合性
"""

import sys
import unicodedata
from collections import defaultdict, Counter
from openpyxl import load_workbook


def nfkc(val):
    if isinstance(val, str):
        return unicodedata.normalize('NFKC', val).strip()
    return val


def safe_str(val):
    if val is None:
        return None
    s = nfkc(str(val))
    return s if s else None


def safe_int(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(float(nfkc(str(val))))
    except (ValueError, TypeError):
        return None


class ValidationReport:
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.info = []

    def error(self, msg):
        self.errors.append(msg)
        print(f'  ❌ {msg}')

    def warn(self, msg):
        self.warnings.append(msg)
        print(f'  ⚠️  {msg}')

    def ok(self, msg):
        self.info.append(msg)
        print(f'  ✅ {msg}')

    def summary(self):
        print(f'\n{"="*60}')
        print(f'検証結果: エラー {len(self.errors)}件 / 警告 {len(self.warnings)}件')
        print(f'{"="*60}')


def check_general_ids(wb_generals, report):
    """武将IDの一意性チェック"""
    print('\n[1] 武将IDの一意性チェック')

    ws = wb_generals['クエリ1']
    ids = []
    names = []

    for row in ws.iter_rows(min_row=3, values_only=True):  # 行3〜（行2=空行）
        if not row or row[0] is None:
            continue
        g_id = safe_int(row[0])
        name = safe_str(row[3])  # 列4: 名前
        rarity = safe_str(row[1])  # 列2: レア
        if g_id is not None:
            ids.append((g_id, name, rarity))

    # ID重複チェック
    id_counts = Counter(item[0] for item in ids)
    duplicates = {k: v for k, v in id_counts.items() if v > 1}

    if duplicates:
        for dup_id, count in duplicates.items():
            dup_generals = [(name, rarity) for gid, name, rarity in ids if gid == dup_id]
            report.error(f'ID重複: {dup_id} ({count}件) → {dup_generals}')
    else:
        report.ok(f'武将ID一意性: OK ({len(ids)}名)')

    return {item[0]: (item[1], item[2]) for item in ids}


def check_skill_links(wb_main, general_data, report):
    """武将技能→技能DBリンクの整合性チェック"""
    print('\n[2] 武将技能→技能DBリンクチェック')

    # 技能DBから全技能名を収集
    ws_skill = wb_main['技能DB']
    skill_names = set()
    for row in ws_skill.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        name = safe_str(row[3])  # 列4: 名称
        if name:
            skill_names.add(name)

    # 武将技能シートから武将の技能名を収集
    ws_busho = wb_main['武将技能']
    busho_skill_names = set()
    for row in ws_busho.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = safe_str(row[3])  # 列4: 名称
        if name:
            busho_skill_names.add(name)

    # 武将技能にあるが技能DBにないもの
    missing = busho_skill_names - skill_names
    if missing:
        for name in sorted(missing):
            report.warn(f'武将技能に存在するが技能DBにない: {name}')
    else:
        report.ok(f'武将技能→技能DBリンク: 全{len(busho_skill_names)}件OK')


def check_treasure_skill_links(wb_main, wb_treasure, report):
    """名宝技能→技能DBリンクの整合性チェック"""
    print('\n[3] 名宝技能→技能DBリンクチェック')

    # 技能DBから全技能名を収集
    ws_skill = wb_main['技能DB']
    skill_names = set()
    for row in ws_skill.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        name = safe_str(row[3])
        if name:
            skill_names.add(name)

    # 名宝一覧の技能名を収集
    ws_list = wb_treasure['名宝一覧']
    treasure_skill_names = set()
    for row in ws_list.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        for i in range(3):
            skill = safe_str(row[8 + i]) if len(row) > 8 + i else None
            if skill and skill != '(技能データなし)':
                treasure_skill_names.add(skill)

    # 名宝シートの技能名
    ws_meiho = wb_main['名宝']
    meiho_skill_names = set()
    for row in ws_meiho.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = safe_str(row[3])
        if name:
            meiho_skill_names.add(name)

    # 名宝一覧にあるが技能DBにない
    missing_from_db = treasure_skill_names - skill_names
    if missing_from_db:
        for name in sorted(missing_from_db):
            report.warn(f'名宝一覧にあるが技能DBにない技能: {name}')

    # 名宝一覧にあるが名宝シートにない
    missing_from_meiho = treasure_skill_names - meiho_skill_names
    if missing_from_meiho:
        for name in sorted(missing_from_meiho):
            report.warn(f'名宝一覧にあるが名宝シートにない技能: {name}')

    if not missing_from_db and not missing_from_meiho:
        report.ok(f'名宝技能リンク: 全{len(treasure_skill_names)}件OK')


def check_treasure_id_collisions(wb_main, report):
    """名宝ID衝突の検出"""
    print('\n[4] 名宝ID衝突チェック')

    ws = wb_main['名宝']
    id_to_weapons = defaultdict(set)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        t_id = safe_int(row[0])
        related = safe_str(row[5])  # 列6: 関連
        if t_id is not None and related:
            id_to_weapons[t_id].add(related)

    collisions = {k: v for k, v in id_to_weapons.items() if len(v) > 1}
    if collisions:
        report.warn(f'名宝ID衝突: {len(collisions)}件')
        for t_id, weapons in sorted(collisions.items()):
            report.warn(f'  ID {t_id}: {" / ".join(sorted(weapons))}')
    else:
        report.ok('名宝ID衝突: なし')


def check_forge_consistency(wb_treasure, report):
    """鍛錬データと名宝一覧の整合性チェック"""
    print('\n[5] 鍛錬データ整合性チェック')

    # 名宝一覧のID集合
    ws_list = wb_treasure['名宝一覧']
    list_ids = set()
    for row in ws_list.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        list_ids.add(safe_int(row[0]))

    # 鍛錬のID集合
    ws_forge = wb_treasure['鍛錬']
    forge_ids = set()
    for row in ws_forge.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        forge_ids.add(safe_int(row[0]))

    # 鍛錬にあるが名宝一覧にないID
    orphan = forge_ids - list_ids
    if orphan:
        for oid in sorted(orphan):
            report.error(f'鍛錬にあるが名宝一覧にないID: {oid}')

    # 名宝一覧にあるが鍛錬にないID
    missing = list_ids - forge_ids
    if missing:
        for mid in sorted(missing):
            report.warn(f'名宝一覧にあるが鍛錬にないID: {mid}')

    if not orphan and not missing:
        report.ok(f'鍛錬↔名宝一覧: 全{len(list_ids)}件一致')
    else:
        report.ok(f'名宝一覧: {len(list_ids)}件, 鍛錬: {len(forge_ids)}件')


def main():
    if len(sys.argv) < 4:
        print('使い方: python validate_data.py <覇道DB.xlsx> <覇道DB_武将_.xlsx> <覇道DB_名宝_.xlsx>')
        sys.exit(1)

    main_path = sys.argv[1]
    generals_path = sys.argv[2]
    treasure_path = sys.argv[3]

    report = ValidationReport()

    print(f'📖 読込中...')
    wb_main = load_workbook(main_path, read_only=True, data_only=True)
    wb_generals = load_workbook(generals_path, read_only=True, data_only=True)
    wb_treasure = load_workbook(treasure_path, read_only=True, data_only=True)

    # 1. 武将ID一意性
    general_data = check_general_ids(wb_generals, report)

    # 2. 武将技能リンク
    check_skill_links(wb_main, general_data, report)

    # 3. 名宝技能リンク
    check_treasure_skill_links(wb_main, wb_treasure, report)

    # 4. 名宝ID衝突
    check_treasure_id_collisions(wb_main, report)

    # 5. 鍛錬整合性
    check_forge_consistency(wb_treasure, report)

    wb_main.close()
    wb_generals.close()
    wb_treasure.close()

    report.summary()


if __name__ == '__main__':
    main()
